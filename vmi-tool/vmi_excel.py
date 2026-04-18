import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font


def largest_remainder_round(amounts: list[float], decimals: int = 2) -> list[float]:
    if not amounts:
        return []
    factor = 10 ** decimals
    target_cents = round(sum(amounts) * factor)
    floors = [int(a * factor) for a in amounts]
    remainder = target_cents - sum(floors)
    fractional = sorted(
        [(a * factor - int(a * factor), i) for i, a in enumerate(amounts)],
        key=lambda x: x[0],
        reverse=True,
    )
    result = floors[:]
    for k in range(remainder):
        result[fractional[k][1]] += 1
    return [r / factor for r in result]


def generate_vmi_excel(account_number: str, rows: list[dict]) -> bytes:
    """
    rows: list of {'rusis': str, 'data': date, 'suma': float, 'valstybe': str}
    Returns xlsx bytes.
    """
    # Group amounts by type for largest-remainder rounding
    type_groups: dict[str, list[tuple[int, float]]] = {}
    for i, row in enumerate(rows):
        t = row["rusis"]
        type_groups.setdefault(t, []).append((i, row["suma"]))

    rounded_sums: list[float] = [0.0] * len(rows)
    for t, entries in type_groups.items():
        indices = [e[0] for e in entries]
        amounts = [e[1] for e in entries]
        for idx, val in zip(indices, largest_remainder_round(amounts)):
            rounded_sums[idx] = val

    wb = Workbook()
    ws = wb.active
    ws.title = "VMI"

    headers = ["saskaita", "rusis", "data", "suma", "valstybe"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10

    for i, row in enumerate(rows):
        ws.append([
            account_number,
            row["rusis"],
            row["data"],
            f"{rounded_sums[i]:.2f}",
            row.get("valstybe", "LT"),
        ])
        date_cell = ws.cell(row=i + 2, column=3)
        date_cell.number_format = "YYYY-MM-DD"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
